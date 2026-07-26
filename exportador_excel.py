"""
exportador_excel.py
=====================
Exportação de todas as fichas técnicas para um arquivo Excel (.xlsx)
com três abas: Resumo, Ingredientes (detalhado) e Custo por Categoria.
"""

from pathlib import Path
from typing import List, Dict

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from models import FichaTecnica

COR_CABECALHO = "B23A2E"  # mesma cor do PDF, sem o "#"


def _formatar_cabecalho(worksheet, n_colunas: int) -> None:
    """Aplica negrito, cor de fundo e auto-ajuste de largura ao cabeçalho."""
    fonte_branca = Font(bold=True, color="FFFFFF")
    preenchimento = PatternFill(start_color=COR_CABECALHO, end_color=COR_CABECALHO, fill_type="solid")
    for col in range(1, n_colunas + 1):
        celula = worksheet.cell(row=1, column=col)
        celula.font = fonte_branca
        celula.fill = preenchimento
        celula.alignment = Alignment(horizontal="center", vertical="center")
    worksheet.freeze_panes = "A2"


def _autoajustar_colunas(worksheet) -> None:
    """Ajusta a largura das colunas conforme o conteúdo (aproximado)."""
    for coluna_celulas in worksheet.columns:
        maior_tamanho = 0
        letra_coluna = get_column_letter(coluna_celulas[0].column)
        for celula in coluna_celulas:
            valor = str(celula.value) if celula.value is not None else ""
            maior_tamanho = max(maior_tamanho, len(valor))
        worksheet.column_dimensions[letra_coluna].width = min(max(maior_tamanho + 3, 10), 45)


def exportar_todas_excel(fichas: List[FichaTecnica], relatorio_categoria: Dict[str, dict], caminho_saida: str) -> str:
    """
    Exporta todas as fichas técnicas para um arquivo .xlsx com 3 abas:
      1. Resumo         -> uma linha por ficha, com todos os indicadores
      2. Ingredientes   -> detalhamento de ingredientes de todas as fichas
      3. Custo por Categoria -> relatório de custo médio por categoria
    Retorna o caminho final do arquivo gerado.
    """
    caminho = Path(caminho_saida)
    caminho.parent.mkdir(parents=True, exist_ok=True)

    if not fichas:
        raise ValueError("Não há fichas técnicas cadastradas para exportar.")

    # ---------- Aba 1: Resumo ----------
    linhas_resumo = []
    for f in fichas:
        linhas_resumo.append({
            "ID": f.id,
            "Nome do Prato": f.nome,
            "Categoria": f.categoria,
            "Porção/Rendimento": f.porcao_rendimento,
            "Tempo de Preparo (min)": f.tempo_preparo_min,
            "Custo Ingredientes (R$)": f.custo_ingredientes,
            "Custo Mão de Obra (R$)": f.custo_mao_de_obra,
            "Custo Total (R$)": f.custo_total_producao,
            "Preço de Venda (R$)": f.preco_venda_final,
            "Margem de Lucro (%)": f.margem_lucro_percentual,
            "Margem de Lucro (R$)": f.margem_lucro_valor,
            "Food Cost (%)": f.food_cost_percentual,
            "Alergênicos": ", ".join(f.alergenicos) if f.alergenicos else "-",
            "Observações": f.observacoes or "-",
            "Atualizado em": f.data_atualizacao,
        })
    df_resumo = pd.DataFrame(linhas_resumo)

    # ---------- Aba 2: Ingredientes detalhados ----------
    linhas_ingredientes = []
    for f in fichas:
        for ing in f.ingredientes:
            linhas_ingredientes.append({
                "Ficha (ID)": f.id,
                "Prato": f.nome,
                "Ingrediente": ing.nome,
                "Quantidade": ing.quantidade,
                "Unidade": ing.unidade,
                "Custo Unitário (R$)": ing.custo_unitario,
                "Custo Total (R$)": ing.custo_total,
            })
    df_ingredientes = pd.DataFrame(linhas_ingredientes) if linhas_ingredientes else pd.DataFrame(
        columns=["Ficha (ID)", "Prato", "Ingrediente", "Quantidade", "Unidade",
                 "Custo Unitário (R$)", "Custo Total (R$)"]
    )

    # ---------- Aba 3: Custo médio por categoria ----------
    linhas_categoria = []
    for categoria, dados in relatorio_categoria.items():
        linhas_categoria.append({
            "Categoria": categoria,
            "Qtd. de Fichas": dados["quantidade_fichas"],
            "Custo Médio (R$)": dados["custo_medio"],
            "Preço Médio (R$)": dados["preco_medio"],
            "Margem Média (%)": dados["margem_media"],
            "Food Cost Médio (%)": dados["food_cost_medio"],
        })
    df_categoria = pd.DataFrame(linhas_categoria)

    # ---------- Gravação do arquivo com formatação ----------
    with pd.ExcelWriter(caminho, engine="openpyxl") as writer:
        df_resumo.to_excel(writer, sheet_name="Resumo", index=False)
        df_ingredientes.to_excel(writer, sheet_name="Ingredientes", index=False)
        df_categoria.to_excel(writer, sheet_name="Custo por Categoria", index=False)

        for nome_aba, df in (
            ("Resumo", df_resumo),
            ("Ingredientes", df_ingredientes),
            ("Custo por Categoria", df_categoria),
        ):
            planilha = writer.sheets[nome_aba]
            _formatar_cabecalho(planilha, n_colunas=max(len(df.columns), 1))
            _autoajustar_colunas(planilha)

    return str(caminho)
